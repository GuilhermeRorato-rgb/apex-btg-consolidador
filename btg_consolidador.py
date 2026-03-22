"""
btg_consolidador.py
====================
Le o extrato BTG Pactual (.xlsx) e preenche a planilha ComDinheiro
(aba 'comdinheiro') com as posicoes do dia.

Uso:
    python btg_consolidador.py <extrato_btg.xlsx> <planilha_cd_vazia.xlsx> [saida.xlsx]

Se [saida.xlsx] nao for informado, gera "saida_comdinheiro_YYYYMMDD.xlsx".
"""

import sys
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import warnings
warnings.filterwarnings("ignore")

NOME_PORTFOLIO = "JBL_Onshore"   # default, sobrescrito em runtime
INSTITUICAO    = "BTG Pactual"
_id_counter = [8657860000]

def next_id():
    _id_counter[0] += 1
    return _id_counter[0]

def to_date(val):
    if val is None or (isinstance(val, float) and pd.isna(val)): return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, pd.Timestamp): return None if pd.isna(val) else val.date()
    # String date
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try: return datetime.strptime(val[:10], "%Y-%m-%d").date()
            except: pass
    return val

def to_float(val):
    if val is None or (isinstance(val, float) and pd.isna(val)): return None
    s = str(val).strip()
    if s in ("-", "", "nan", "NaN", "None"): return 0.0
    try: return float(s)
    except: return None

def parse_taxa(taxa_str):
    taxa_str = str(taxa_str).strip()
    indexador, percent, taxa_pre, gross_up = None, None, None, None
    if "CDI" in taxa_str.upper():
        indexador = "CDI"
        m = re.search(r'([\d,\.]+)%', taxa_str)
        if m: percent = float(m.group(1).replace(",", "."))
        gross_up = 117.647059
    elif "IPCA" in taxa_str.upper():
        indexador = "ipca"
        m = re.search(r'[\+\s]*([\d,\.]+)%', taxa_str)
        if m: taxa_pre = float(m.group(1).replace(",", "."))
    elif "%" in taxa_str:
        indexador = "prefix"; percent = 0.0
        m = re.search(r'([\d,\.]+)%', taxa_str)
        if m: taxa_pre = float(m.group(1).replace(",", "."))
    return indexador, percent, taxa_pre, gross_up

def build_apelido(tipo, code, vencimento, indexador, percent, taxa_pre):
    if vencimento is None: return None
    vc_fmt = vencimento.strftime("%d/%m/%Y") if hasattr(vencimento, 'strftime') else str(vencimento)
    if tipo == "titulo":
        taxa_desc = f" IPCA+{taxa_pre:.2f}%" if taxa_pre else ""
        return f"NTNB {vc_fmt}{taxa_desc}"
    elif tipo in ("cra", "cri"):
        taxa_desc = (f" {percent:.0f}% CDI" if indexador=="CDI" and percent
                     else (f" {taxa_pre:.2f}%" if taxa_pre else ""))
        return f"{tipo.upper()} BTG Pactual {vc_fmt} {code}{taxa_desc}"
    return None

TIPO_MAP = {"CDB":"cdb","CRA":"cra","CRI":"cri","NTNB":"titulo","NTNF":"titulo",
            "LCI":"lci","LCA":"lca","DEBENTURE":"debenture"}

# ─────────────────────────────────────────────────────
# PARSER: FUNDOS
# ─────────────────────────────────────────────────────

def parse_fundos(df_raw):
    rows = []
    col_b = df_raw.iloc[:, 1]
    in_fundos = in_ca = False
    current_cnpj = current_name = None

    for i, val in enumerate(col_b):
        val_str = str(val).strip() if pd.notna(val) else ""

        if "Posição > Portfólio de fundos" in val_str:
            in_fundos, in_ca = True, False; continue
        if "Portfólio de Carteiras Administradas" in val_str:
            in_ca, in_fundos = True, False; continue
        if any(k in val_str for k in ("Detalhamento","Rentabilidade","Movimentações")):
            break

        if not (in_fundos or in_ca): continue

        cnpj_match = re.search(r'CNPJ:\s*([\d./\-]+)', val_str)
        if cnpj_match:
            current_cnpj = cnpj_match.group(1); current_name = val_str; continue

        if in_ca and val_str and val_str not in ("NaN",""):
            next_raw = df_raw.iloc[i+1, 1] if i+1 < len(df_raw) else None
            if isinstance(next_raw, datetime):
                current_cnpj = val_str; current_name = val_str; continue

        raw_val = df_raw.iloc[i, 1]
        if not isinstance(raw_val, datetime) or not current_cnpj: continue
        data_op = raw_val.date()

        if in_fundos:
            quantidade  = to_float(df_raw.iloc[i, 3])
            preco_unit  = to_float(df_raw.iloc[i, 4])
            total_bruto = to_float(df_raw.iloc[i, 5])
            ir_val      = to_float(df_raw.iloc[i, 6]) or 0.0
            iof_val     = to_float(df_raw.iloc[i, 7]) or 0.0

            cnpj_ativo = current_cnpj
            if "Subclasse" in current_name:
                subm = re.search(r'Cód\. Subclasse:\s*(\S+)', current_name)
                cnpj_ativo = current_cnpj + ("_subclasse3" if subm else "_unica")
            elif re.search(r'\(E\d+\)', current_name):
                cnpj_ativo = current_cnpj + "_unica"

            rows.append({"nome_portfolio":NOME_PORTFOLIO,"ativo":cnpj_ativo,"tipo_ativo":"fundo",
                "data_operacao":data_op,"id":next_id(),"id2":None,"CV":"C",
                "data_cotizacao":data_op,"data_liquidacao":data_op,
                "preco_unitario":preco_unit,"quantidade":quantidade,"total_bruto":total_bruto,
                "alt_caixa":0,"data_vencimento":None,"custo_transacao":None,"flag_tribut":None,
                "marcacao":None,"indexador":None,"percent":None,"taxa_pre":None,"gross_up":None,
                "taxa_cupom_pre":None,"instituicao_financeira":INSTITUICAO,"banco":None,
                "flag_liquidez":None,"tempo_liquidez":None,"data_liquidez":None,"apelido":None,
                "flag_provisao":None,"IR":ir_val,"IOF":iof_val,"total_liquido":None,"campo_preciso":None})

        elif in_ca:
            quantidade  = to_float(df_raw.iloc[i, 2])
            preco_unit  = to_float(df_raw.iloc[i, 3])
            total_bruto = to_float(df_raw.iloc[i, 4])
            rows.append({"nome_portfolio":NOME_PORTFOLIO,"ativo": NOME_PORTFOLIO.split("_")[0] + "_CART_RF","tipo_ativo":"generico",
                "data_operacao":data_op,"id":next_id(),"id2":None,"CV":"C",
                "data_cotizacao":None,"data_liquidacao":data_op,
                "preco_unitario":preco_unit,"quantidade":quantidade,"total_bruto":total_bruto,
                "alt_caixa":0,"data_vencimento":None,"custo_transacao":None,"flag_tribut":0.0,
                "marcacao":None,"indexador":None,"percent":None,"taxa_pre":None,"gross_up":None,
                "taxa_cupom_pre":None,"instituicao_financeira":INSTITUICAO,"banco":None,
                "flag_liquidez":5.0,"tempo_liquidez":None,"data_liquidez":None,
                "apelido":"Carteira Administrada RF","flag_provisao":None,
                "IR":None,"IOF":None,"total_liquido":None,"campo_preciso":None})
    return rows


# ─────────────────────────────────────────────────────
# PARSER: RENDA FIXA
# Estrutura da aba (colunas 0-indexed a partir de B=1):
#   Secao Posicoes:
#     B=Emissor, C=Ativo, D=Emissao, E=Vencimento, F=Liquidez,
#     G=DiasCarencia, H=DataInicLiq, I=Taxa, J=Qtd, K=Preco, L=SaldoBruto, M=IR, N=IOF, O=SaldoLiq
#   Secao Detalhamento (B=Ativo, dados nas colunas seguintes):
#     B=Ativo(codigo), C=Emissao, D=Vencimento, E=Aquisicao, F=Liquidez,
#     G=DiasCarencia, H=DataInicLiq, I=TaxaCompra, J=Qtd, K=PrecoCompra, L=ValorCompra,
#     M=PrecoAtual, N=SaldoBruto, O=IR, P=IOF, Q=SaldoLiq
# ─────────────────────────────────────────────────────

def parse_renda_fixa(df_raw):
    rows = []
    col_b = df_raw.iloc[:, 1]
    n = len(df_raw)

    # Coletar dados resumidos de posicoes (emissor por ativo)
    posicoes = {}
    current_type = None
    in_posicoes = False

    for i, val in enumerate(col_b):
        val_str = str(val).strip() if pd.notna(val) else ""
        if "Posições Detalhadas" in val_str: break
        if val_str == "Posições": in_posicoes = True; continue
        if not in_posicoes: continue

        pm = re.match(r"Posição > (\w+)", val_str)
        if pm: current_type = TIPO_MAP.get(pm.group(1).upper(), pm.group(1).lower()); continue

        if val_str in ("Emissor","Total","","NaN") or not current_type: continue

        emissor    = val_str
        ativo_code = str(df_raw.iloc[i, 2]).strip() if pd.notna(df_raw.iloc[i, 2]) else ""
        if not ativo_code or ativo_code in ("NaN","nan","","Total"): continue

        try:
            ir_v  = to_float(df_raw.iloc[i, 12]) or 0.0
            iof_v = to_float(df_raw.iloc[i, 13]) or 0.0
        except: ir_v = iof_v = 0.0

        code_clean = re.sub(r'^(CRA|CRI|CDB|NTNB|NTNF)-', '', ativo_code)
        posicoes[(current_type, code_clean)] = {"emissor": emissor, "ir": ir_v, "iof": iof_v}

    # Processar detalhamento
    # Layout detalhamento (col indices, 0-based from column A):
    #   1=Ativo, 2=Emissao, 3=Vencimento, 4=Aquisicao, 5=Liquidez,
    #   6=DiasCarencia, 7=DataInicLiq, 8=TaxaCompra, 9=Qtd, 10=PrecoCompra,
    #   11=ValorCompra, 12=PrecoAtual, 13=SaldoBruto, 14=IR, 15=IOF, 16=SaldoLiq
    in_det = False
    current_type_det = current_code_det = current_emissor = None

    for i, val in enumerate(col_b):
        val_str = str(val).strip() if pd.notna(val) else ""

        if "Posições Detalhadas" in val_str: in_det = True; continue
        if not in_det: continue
        if any(k in val_str for k in ("Posição Consolidada","Movimentações")): break

        # Cabecalho de detalhamento: "Detalhamento > CDB | BANCO MASTER S/A"
        dm = re.match(r"Detalhamento > (\w+)(?:\s*\|\s*(.+))?$", val_str)
        if dm:
            current_type_det = TIPO_MAP.get(dm.group(1).upper(), dm.group(1).lower())
            current_emissor  = dm.group(2).strip() if dm.group(2) else None
            current_code_det = None
            continue

        if val_str in ("Ativo","Total","","NaN"): continue

        # Linha de codigo do ativo (a primeira linha nao-header, nao-datetime)
        # O codigo fica na propria coluna B; na mesma linha as colunas C..Q tem os dados
        if current_type_det and current_code_det is None:
            # Esta linha e o codigo do ativo (ex: "CRA-CRA02300209")
            code_raw = val_str
            # Checar se colunas de dados estao nesta linha
            try:
                emissao_check = df_raw.iloc[i, 2]
            except: emissao_check = None

            if emissao_check is not None and not (isinstance(emissao_check, float) and pd.isna(emissao_check)):
                # E uma linha de dados (codigo + dados na mesma linha)
                current_code_det = re.sub(r'^(CRA|CRI|CDB|NTNB|NTNF)-', '', code_raw).strip()
                # Processar esta linha diretamente
                _process_det_row(df_raw, i, current_type_det, current_code_det, current_emissor, posicoes, rows)
                continue
            else:
                # So o codigo, dados vem na proxima linha
                current_code_det = re.sub(r'^(CRA|CRI|CDB|NTNB|NTNF)-', '', code_raw).strip()
                continue

        # Linha de dados depois do codigo
        if current_type_det and current_code_det:
            emissao_check = df_raw.iloc[i, 2] if i < n else None
            if emissao_check is not None and not (isinstance(emissao_check, float) and pd.isna(emissao_check)):
                _process_det_row(df_raw, i, current_type_det, current_code_det, current_emissor, posicoes, rows)

    return rows


def _process_det_row(df_raw, i, tipo, code, emissor_det, posicoes, rows):
    """Processa uma linha de dados do detalhamento de renda fixa."""
    try:
        ativo_code_raw = str(df_raw.iloc[i, 1]).strip()
        emissao_raw    = df_raw.iloc[i, 2]
        vencimento_raw = df_raw.iloc[i, 3]
        aquisicao_raw  = df_raw.iloc[i, 4]
        taxa_compra    = df_raw.iloc[i, 8]
        quantidade_s   = df_raw.iloc[i, 9]
        preco_compra_s = df_raw.iloc[i, 10]
        valor_compra_s = df_raw.iloc[i, 11]
        preco_atual_s  = df_raw.iloc[i, 12]
        total_bruto_s  = df_raw.iloc[i, 13]
        ir_s           = df_raw.iloc[i, 14]
        iof_s          = df_raw.iloc[i, 15]
        saldo_liq_s    = df_raw.iloc[i, 16]
    except IndexError:
        return

    emissao    = to_date(emissao_raw)
    vencimento = to_date(vencimento_raw)
    aquisicao  = to_date(aquisicao_raw)
    quantidade = to_float(quantidade_s)
    preco_compra = to_float(preco_compra_s)
    valor_compra = to_float(valor_compra_s)
    ir_d   = to_float(ir_s)  or 0.0
    iof_d  = to_float(iof_s) or 0.0
    saldo_liq = to_float(saldo_liq_s)

    if quantidade is None: return

    data_op  = aquisicao or emissao
    venc_use = vencimento

    indexador, percent, taxa_pre, gross_up = parse_taxa(str(taxa_compra))
    pos_info = posicoes.get((tipo, code), {})

    # Montar ativo_id
    if tipo == "titulo":
        em_str = emissao.strftime("%d%m%Y") if emissao else "00000000"
        vc_str = venc_use.strftime("%d%m%Y") if venc_use else "00000000"
        ativo_id = f"NTNB_{em_str}_{vc_str}"
    elif tipo in ("cra","cri"):
        ativo_id = f"cetip_{code}"
    else:
        # CDB
        emissor  = emissor_det or pos_info.get("emissor","")
        vc_fmt   = venc_use.strftime("%d/%m/%Y") if venc_use else ""
        m = re.search(r'([\d,\.]+)%', str(taxa_compra))
        taxa_pct = m.group(1).replace(",",".") if m else ""
        ativo_id = f"CDB {emissor} {vc_fmt} {taxa_pct}% (em liquidacao)" if taxa_pct else f"CDB {emissor} {vc_fmt} (em liquidacao)"

    # Para CDB: preco=1, qtd=saldo liquido
    if tipo == "cdb":
        preco_unit_use = 1.0
        qtd_use  = saldo_liq if saldo_liq else (to_float(total_bruto_s) or 0) - ir_d - iof_d
        total_use = qtd_use
    else:
        preco_unit_use = preco_compra
        qtd_use  = quantidade
        total_use = valor_compra

    rows.append({
        "nome_portfolio": NOME_PORTFOLIO,
        "ativo":          ativo_id,
        "tipo_ativo":     tipo,
        "data_operacao":  data_op,
        "id":             next_id(),
        "id2":            None,
        "CV":             "C",
        "data_cotizacao": None,
        "data_liquidacao":data_op,
        "preco_unitario": preco_unit_use,
        "quantidade":     qtd_use,
        "total_bruto":    total_use,
        "alt_caixa":      0,
        "data_vencimento":venc_use,
        "custo_transacao":None,
        "flag_tribut":    0.0 if tipo in ("cdb","cra","cri") else None,
        "marcacao":       "curva" if tipo in ("cra","cri","titulo") else None,
        "indexador":      indexador,
        "percent":        percent,
        "taxa_pre":       taxa_pre,
        "gross_up":       gross_up,
        "taxa_cupom_pre": None,
        "instituicao_financeira": INSTITUICAO,
        "banco":          (emissor_det or pos_info.get("emissor")) if tipo=="cdb" else None,
        "flag_liquidez":  0.0 if tipo=="cdb" else None,
        "tempo_liquidez": None,
        "data_liquidez":  None,
        "apelido":        build_apelido(tipo, code, venc_use, indexador, percent, taxa_pre),
        "flag_provisao":  None,
        "IR":             ir_d,
        "IOF":            iof_d,
        "total_liquido":  saldo_liq,
        "campo_preciso":  None,
    })


# ─────────────────────────────────────────────────────
COLUNAS_ORDEM = [
    "nome_portfolio","ativo","tipo_ativo","data_operacao","id","id2","CV",
    "data_cotizacao","data_liquidacao","preco_unitario","quantidade","total_bruto",
    "alt_caixa","data_vencimento","custo_transacao","flag_tribut","marcacao",
    "indexador","percent","taxa_pre","gross_up","taxa_cupom_pre",
    "instituicao_financeira","banco","flag_liquidez","tempo_liquidez","data_liquidez",
    "apelido","flag_provisao","IR","IOF","total_liquido","campo_preciso",
]


def write_to_template(df, template_path, output_path):
    wb = load_workbook(template_path)
    ws = wb["comdinheiro"]

    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "nome_portfolio":
                header_row = cell.row; break
        if header_row: break
    if not header_row: header_row = 1

    col_map = {cell.value: cell.column for cell in ws[header_row] if cell.value in COLUNAS_ORDEM}

    for r in range(header_row + 1, ws.max_row + 1):
        for c in ws[r]: c.value = None

    for idx, (_, row_data) in enumerate(df.iterrows()):
        excel_row = header_row + 1 + idx
        for col_name, col_idx in col_map.items():
            val = row_data.get(col_name)
            cell = ws.cell(row=excel_row, column=col_idx)
            if val is None or (isinstance(val, float) and pd.isna(val)):
                cell.value = None
            elif isinstance(val, pd.Timestamp):
                cell.value = None if pd.isna(val) else val.to_pydatetime()
            else:
                cell.value = val

    wb.save(output_path)
    print(f"Arquivo salvo: {output_path}")
    print(f"Total de linhas: {len(df)}")


def consolidar(extrato_path, template_path, output_path=None, nome_portfolio=None):
    global NOME_PORTFOLIO
    if nome_portfolio and nome_portfolio.strip():
        NOME_PORTFOLIO = nome_portfolio.strip()
    if output_path is None:
        output_path = f"saida_comdinheiro_{datetime.today().strftime('%Y%m%d')}.xlsx"

    print(f"Lendo extrato: {extrato_path}")
    extrato = pd.read_excel(extrato_path, sheet_name=None, header=None)

    all_rows = []
    if "Renda Fixa" in extrato:
        print("Processando Renda Fixa...")
        rf = parse_renda_fixa(extrato["Renda Fixa"])
        print(f"  {len(rf)} posicoes")
        all_rows.extend(rf)

    if "Fundos" in extrato:
        print("Processando Fundos...")
        fundos = parse_fundos(extrato["Fundos"])
        print(f"  {len(fundos)} posicoes")
        all_rows.extend(fundos)

    print(f"Total: {len(all_rows)} posicoes")
    df = pd.DataFrame(all_rows, columns=COLUNAS_ORDEM)
    write_to_template(df, template_path, output_path)
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__); sys.exit(1)
    consolidar(sys.argv[1], sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else None)
